import os
import urllib.request
import ssl
import openpyxl
import re



def download_file(url, folder_path, file_name):

    # Create a folder if it does not exist
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Full path to file
    file_path = os.path.join(folder_path, file_name)

    # Create a secure SSL context
    context = ssl.create_default_context()
    context.check_hostname = False
    context.verify_mode = ssl.CERT_NONE

    # Download file with secure SSL context
    with urllib.request.urlopen(url, context=context) as u, open(file_path, 'wb') as f:
        f.write(u.read())
    print(f"File {file_name} downloaded and saved in folder {folder_path}")
    return


def read_zip_band_from_file(file_path, sheet_name):

    workbook = openpyxl.load_workbook(file_path, data_only=True)  # data_only=True - получение значений, а не формул
    sheet = workbook[sheet_name]

    #Get all rows
    rows = sheet.iter_rows()

    # {zip_start: [zip_start, zip_end], ...} = '00500': ['00500', '00599']
    zip_band_dict = {row[0].value.split("-")[0]: row[0].value.split("-") for row in rows if row[0].value}

    # [[zip_start, zip_end], [zip_start, zip_end], ...] = ['00500', '00599'], ['01000', '01099'], ['01100', '01199'],
    zip_band_list = [value for key, value in zip_band_dict.items()]

    zip_band_list.pop(0) # Delete first element - it is a header

    print(zip_band_list)
    print()

    return zip_band_list, zip_band_dict


def get_reference_range(file_path):

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Get the 5th row, where the zip code range is specified for verification
    row = sheet[5]

    row_text = [str(cell.value) for cell in row]

    # Search for numbers in the text of the line
    # Convert to string in format 01234
    pattern = r'\d{3}-\d{2}'
    matches = re.findall(pattern, ' '.join(row_text))
    if len(matches) == 2:
        ref_start = matches[0].replace("-", "")  # Удаляем символ "-" 012-34 -> 01234 Референсный диапазон zip кодов
        ref_end = matches[1].replace("-", "")

        print(f"Reference range is {ref_start} {ref_end} for {file_path}")
    else:
        print("ERROR - Find incorrect number of numbers in the line")

    return ref_start, ref_end


def download_all_files(zip_band_list, url, folder_path, count_files=None):
    count = 0
    index = 0
    while index < len(zip_band_list[:]):# skip first element - it is a header
        zip_band = zip_band_list[index]
        zip_start = zip_band[0]
        zip_end = zip_band[1]

        name = zip_start[:-2]  # Delete last 2 symbols for download file for example 011.xls

        url_file = url + f'{name}.xls'  # url_file = f'https://www.ups.com/media/us/currentrates/zone-csv/{name}.xls'
        file_name = f'{name}.xlsx'  # Rename file to *.xlsx

        download_file(url_file, folder_path, file_name)
        check, ref_start, ref_end = check_zip_code_from_load_file(folder_path, file_name, zip_start, zip_end)
        if check == False:
            expand_zip_band_list(zip_band_list, index, ref_start, ref_end, zip_start, zip_end)
            # print(index, zip_band_list[index], zip_band_list[index + 1], zip_band_list[index + 2])


        index += 1

        # limit download files
        count += 1
        if count == count_files:
            print(f"--- WARNING --- Limit download files = {count_files}")
            break
        # ----------------------------
    return


def expand_zip_band_list(zip_band_list, index, ref_start, ref_end, zip_start, zip_end):
    len_r_start = len(ref_start) - len(str(int(ref_start)))
    len_r_end = len(ref_end) - len(str(int(ref_end)))
    len_z_end = len(zip_end) - len(str(int(zip_end)))

    # convert to string in format 01234
    zip_band_list[index ] = ["0" * len_r_start + str(int(ref_start) - 1), "0" * len_r_end + str(int(ref_end))]

    # ВНИМАНИЕ! Возможно зацикливание, с бесконечным расширением диапазона
    # расширяем диапазон zip кодов, ТОЛЬКО если zip_end > ref_конец
    # WARNING! Infinite expansion of the range is possible
    # expand the range of zip codes ONLY if zip_end > ref_end

    if int(zip_end)>int(ref_end):
        zip_band_list.insert(index + 1, ["0" * len_r_end + str(int(ref_end) + 1), "0" * len_z_end + str(int(zip_end))])

    return


def check_zip_code_from_load_file(folder_path, file_name, zip_start, zip_end):
    # Чтение полученого файла Excel и поиск диапазона чисел
    file_path = folder_path + '/' + file_name
    ref_start, ref_end = get_reference_range(file_path)
    # 00501-1 <= 00500 <= 00599 and 00599 >= 00599
    if int(ref_start) - 1 <= int(zip_start) <= int(ref_end) and int(zip_end) == int(ref_end):

        print(f'Diapason is TRUE for {file_name}')
        print("---------------------------------")
        answer = True
    else:
        print(" ---- WARNING - Diapason is FALSE ---- ")
        print(f'Diapason getting from user {zip_start} - {zip_end}')
        print(f'Diapason is FALSE for {file_name} {ref_start} - {ref_end}')
        print("---------------------------------")
        answer = False
    return answer, ref_start, ref_end


def write_data_to_txt(file_name, zip_band_list):
    with open(file_name, "w") as f:
        f.write(f"UPS zone ranges,	zip from,	zip to\n") # Header
        for item in zip_band_list:
            f.write(f"{item[0]}-{item[1]},{item[0]}, {item[1]}\n")


def write_to_excel(file_path, zip_band_list):
    # Create new Excel file and add a worksheet.
    wb = openpyxl.Workbook()
    sheet = wb.active  # New sheet
    sheet.title = "UPS zip ranges"
    sheet.append(["UPS zone ranges", "zip from", "zip to"])  # Header

    # Save elements to Excel cells
    for row in zip_band_list:
        row_xlsx = [f"{row[0]}-{row[1]}", row[0], row[1]]
        sheet.append(row_xlsx)

    wb.save(file_path)
    print("=============================")
    print(f"The file [{file_path}] saved!")
    return


if __name__ == '__main__':
    ## 1) Read the incoming Excel file to get the zip code ranges
    file_path = 'Inbox Data/Carriers zone ranges.xlsx'
    sheet_name = 'UPS zip ranges'
    zip_band_list, zip_band_dict = read_zip_band_from_file(file_path, sheet_name)

    ## 2) Download files from the site and check the zip code ranges from the downloaded files
    url = r'https://www.ups.com/media/us/currentrates/zone-csv/'
    folder_output = 'Output Data'

    # COUNT_FILES = None for all files, for test = 20
    COUNT_FILES = 10

    download_all_files(zip_band_list, url, folder_output, COUNT_FILES)

    ## 3) Save to .xlsx file the corrected range.

    file_xlsx = f"{folder_output}/NEW Carriers zone ranges.xlsx"
    write_to_excel(file_xlsx, zip_band_list)

    ## 4) Save .txt file. (For ease of checking in the IDE write to txt)
    file_txt = f"{folder_output}/output.txt"
    write_data_to_txt(file_txt, zip_band_list)

    # print(zip_band_list)
